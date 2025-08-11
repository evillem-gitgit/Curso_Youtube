"""
🔐 1. Sistema de Login Simples
Contexto:
Você está criando um sistema básico de autenticação de usuários.

Desafio:
Crie um programa que:

Armazene usuários com nome e senha.

Permita um login pedindo nome e senha.

Verifique se o login está correto e exiba uma mensagem apropriada.
"""



def verificar_cadastro_e_cadastrar(nome, senha):
    from openpyxl import load_workbook

        # Abre a planilha existente
    wb = load_workbook('Exercitando_CursoDoYoutube/Estudando.xlsx')

    planilha = wb['Planilha1']

    nome = str(input('Digite seu nome: ').title().strip())
    senha = str(input('Digite sua senha: '))


    cadastrado = False
    for linha in planilha.iter_rows(min_row=2, values_only=True):
        
        nomeNaPlanilha, senhaNaPlanilha = linha

        if nomeNaPlanilha == nome and senhaNaPlanilha == senha:
            cadastrado = True
            break

    msg = 'Você já possui cadastro!' if cadastrado 'Você ainda não possui o cadastro!'
        print()
    
    else:
        print()
        print()
        realizar_cadastro('nome', 'senha')    











def realizar_cadastro(nome, senha):
    from openpyxl import load_workbook

    wb = load_workbook('Exercitando_CursoDoYoutube/Estudando.xlsx')

    planilha = wb['Planilha1']
    
    print('Você agora está realizando o cadastro!')

    nome = str(input('Digite seu nome: ').title().strip())
    senha = str(input('Digite sua senha: '))

    
    #title() deixa as 1º letras em maiusculo
    #strip() remove espaços vazios no começo e fim da str

    planilha.append([nome, senha])
    wb.save('Exercitando_CursoDoYoutube/Estudando.xlsx')#salva as alterações
    print()
    print("Cadastro realizado com sucesso!")

  
        











def realizar_login(nome, senha):
        nome = str(input('Digite seu nome: ').title().strip())
        senha = str(input('Digite sua senha: '))


        loginOk = False

        for linha in planilha.iter_rows(min_row=2, values_only=True):
            nomeNaPlanilha, senhaNaPlanilha = linha

            if nome == nomeNaPlanilha and senha == senhaNaPlanilha:
                loginOk = True

        if loginOk:
            print('✅ Login bem-sucedido!')

        else:
            print('❌ Nome ou senha incorreto!')
            print()
    











# Q1a
from openpyxl import load_workbook

# Abre a planilha existente
wb = load_workbook('Exercitando_CursoDoYoutube/Estudando.xlsx')


planilha = wb['Planilha1']


while True:
    print()
    print("Cadastresse ou realize o login!")
    print("Escolha 1 para: cadastro")
    print("Escolha 2 para: login")
    print()
    escolha = int(input('Digite sua escolha aqui: '))
    

    if escolha == 1:

        verificar_cadastro_e_cadastrar('nome', 'senha')



    if escolha == 2:

        realizar_login('nome', 'senha')







# logged_user = True
# msg = 'usuario logado' if logged_user else 'usuario precisa logar'

# print(msg)

# idade = input('qual a sua idade: ')

# if not idade.isnumeric():
#     print('digite so numeros')

# else:
#     idade = int(idade)
#     de_maior = (idade >= 18)

#     msg = 'pode acessar' if de_maior else 'não pode acessar n vey'
# 



